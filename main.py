# This is a sample Python script.

# Press Shift+F10 to execute it or replace it with your code.
# Press Double Shift to search everywhere for classes, files, tool windows, actions, and settings.
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
import PySimpleGUI as sg
import openpyxl
import datetime
import xlsxwriter
from Safetycopy import *


def simple_read_commands():
    wb = load_workbook(filename='test_book.xlsx')
    sheet_ranges = wb['range names']
    print(sheet_ranges['D18'].value)


def simple_write_commands():
    wb = Workbook()
    dest_filename = 'test_book.xlsx'
    ws1 = wb.active
    ws1.title = "range names"

    for row in range(1, 40):
        ws1.append(range(600))

    ws1.merge_cells('A2:D3')
    ws1.unmerge_cells('A2:D3')
    ws1.insert_rows(7)

    ws2 = wb.create_sheet(title="PI")
    ws2['F5'] = 3.14

    ws3 = wb.create_sheet(title="Data")
    for row in range(10, 20):
        for col in range(27, 54):
            _ = ws3.cell(column=col, row=row, value="{0}".format(get_column_letter(col)))

    print(ws3['AA10'].value)

    ws4 = wb.create_sheet(title="Fold")
    ws4.column_dimensions.group('A', 'D', hidden=True)
    ws4.row_dimensions.group(1, 10, hidden=True)

    wb.save(filename=dest_filename)


def same_transfer(source, destination):
    row_start = int(input("row_start:"))
    row_end = int(input("row_end:"))
    column_start = int(input("column_start:"))
    column_end = int(input("column_end:"))
    print("rows", row_start, "-", row_end, "columns", column_start, "-", column_end)

    wb1 = load_workbook(filename=source)
    wb2 = load_workbook(filename=destination)
    ws1 = wb1.active
    ws2 = wb2.active
    for row in range(row_start, row_end + 1):
        for col in range(column_start, column_end + 1):
            c = ws1.cell(row=row, column=col)
            ws2.cell(row=row, column=col).value = c.value
    wb2.save(destination)
    return


def different_transfer(source, destination):
    row_start_source = int(input("row_start:"))
    row_end_source = int(input("row_end:"))
    column_start_source = int(input("column_start:"))
    column_end_source = int(input("column_end:"))
    print("rows in source file", row_start_source, "-", row_end_source, "columns in source file", column_start_source,
          "-", column_end_source)
    row_start_destination = int(input("row_start:"))
    row_end_destination = int(input("row_end:"))
    column_start_destination = int(input("column_start:"))
    column_end_destination = int(input("column_end:"))
    print("rows in destination file", row_start_destination, "-", row_end_destination,
          "columns in destination file", column_start_destination, "-", column_end_destination)
    wb1 = load_workbook(filename=source)
    wb2 = load_workbook(filename=destination)
    ws1 = wb1.active
    ws2 = wb2.active
    c = []
    for row in range(row_start_source, row_end_source + 1):
        for col in range(column_start_source, column_end_source + 1):
            c.append(ws1.cell(row=row, column=col))
    print(c)

    return


def actual_transfer(source, destination):
    wb1 = load_workbook(filename=source)
    wb2 = load_workbook(filename=destination)
    ws1 = wb1.active
    ws2 = wb2.active
    to_append = ()

    #used for converting the column letter to numbers so i get an easier coordinate to work with
    counter = 0
    for column in range(1, ws1.max_column):
        column_letter = get_column_letter(column)
        for row in range(1, ws1.max_row):
            counter += 1
            ws1[column_letter+str(row)] = counter




def copy_sheet(source, target):
    wb = Workbook
    source = wb.active
    target = wb.copy_worksheet(source)
    return


if __name__ == '__main__':


    simple_write_commands()
    #simple_read_commands()

    #same_transfer(values[0], values[1])
    #different_transfer(values[0], values[1])
    #move(values[0], values[1])
# See PyCharm help at https://www.jetbrains.com/help/pycharm/
